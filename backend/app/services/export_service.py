from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus.flowables import HRFlowable
from reportlab.graphics.shapes import Drawing, String, Line, Rect
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics.widgets.markers import makeMarker
from datetime import datetime
from typing import List, Dict, Any, Optional
import os
import csv

# SupportX Brand Colors
SUPPORTX_PURPLE = colors.HexColor('#8b5cf6')
SUPPORTX_PURPLE_DARK = colors.HexColor('#6b21a8')
SUPPORTX_PURPLE_LIGHT = colors.HexColor('#f5f3ff')
SUPPORTX_GREEN = colors.HexColor('#10b981')
SUPPORTX_RED = colors.HexColor('#ef4444')
SUPPORTX_BLUE = colors.HexColor('#3b82f6')
SUPPORTX_YELLOW = colors.HexColor('#f59e0b')
SUPPORTX_GRAY = colors.HexColor('#6b7280')

# Chart color palette
CHART_COLORS = [
    colors.HexColor('#8b5cf6'),  # Purple
    colors.HexColor('#10b981'),  # Green
    colors.HexColor('#3b82f6'),  # Blue
    colors.HexColor('#f59e0b'),  # Yellow
    colors.HexColor('#ef4444'),  # Red
    colors.HexColor('#6366f1'),  # Indigo
    colors.HexColor('#ec4899'),  # Pink
    colors.HexColor('#14b8a6'),  # Teal
]

# Excel colors
EXCEL_PURPLE = "8B5CF6"
EXCEL_PURPLE_LIGHT = "F5F3FF"
EXCEL_GREEN = "10B981"
EXCEL_RED = "EF4444"


class ExportService:
    """Service for exporting reports to Excel, PDF, and CSV with professional styling and charts"""

    def __init__(self, db=None):
        self.db = db
        self.export_dir = "exports/reports"
        os.makedirs(self.export_dir, exist_ok=True)
        self.logo_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "static", "logo.png")

    # ============================================================================
    # MAIN EXPORT METHODS
    # ============================================================================

    def export_to_excel(self, data: Any, report_name: str, sheet_name: str = "Report") -> str:
        """Export report data to Excel with charts"""
        filename = f"SupportX_{report_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Styles
        header_fill = PatternFill(start_color=EXCEL_PURPLE, end_color=EXCEL_PURPLE, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=18, color="6B21A8")
        subtitle_font = Font(bold=True, size=12, color="6B7280")
        alt_fill = PatternFill(start_color=EXCEL_PURPLE_LIGHT, end_color=EXCEL_PURPLE_LIGHT, fill_type="solid")
        border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        if report_name == "sla_compliance":
            self._export_sla_to_excel(wb, data, header_fill, header_font, title_font, subtitle_font, alt_fill, border)
        elif report_name == "ticket_aging":
            self._export_aging_to_excel(wb, data, header_fill, header_font, title_font, subtitle_font, alt_fill, border)
        elif report_name == "technician_performance":
            self._export_performance_to_excel(wb, data, header_fill, header_font, title_font, subtitle_font, alt_fill, border)
        elif report_name == "ticket_volume":
            self._export_volume_to_excel(wb, data, header_fill, header_font, title_font, subtitle_font, alt_fill, border)
        elif report_name == "category_breakdown":
            self._export_category_to_excel(wb, data, header_fill, header_font, title_font, subtitle_font, alt_fill, border)
        else:
            self._export_generic_to_excel(ws, data, header_fill, header_font, alt_fill, border)

        wb.save(filepath)
        return filepath

    def export_to_pdf(self, data: Any, report_name: str, title: str) -> str:
        """Export report data to PDF with charts and professional styling"""
        filename = f"SupportX_{report_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.export_dir, filename)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=letter,
            topMargin=80,
            bottomMargin=45,
            leftMargin=40,
            rightMargin=40
        )

        story = []

        if report_name == "sla_compliance":
            story = self._build_sla_pdf(data, title)
        elif report_name == "ticket_aging":
            story = self._build_aging_pdf(data, title)
        elif report_name == "technician_performance":
            story = self._build_performance_pdf(data, title)
        elif report_name == "ticket_volume":
            story = self._build_volume_pdf(data, title)
        elif report_name == "category_breakdown":
            story = self._build_category_pdf(data, title)
        else:
            story = self._build_generic_pdf(data, title)

        def add_page_elements(canvas, doc):
            self._create_header(canvas, doc, title)
            self._create_footer(canvas, doc)

        doc.build(story, onFirstPage=add_page_elements, onLaterPages=add_page_elements)
        return filepath

    def export_to_csv(self, data: Any, report_name: str) -> str:
        """Export report data to CSV"""
        filename = f"SupportX_{report_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(self.export_dir, filename)

        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)

            if report_name == "sla_compliance":
                self._export_sla_to_csv(writer, data)
            elif report_name == "ticket_aging":
                self._export_aging_to_csv(writer, data)
            elif report_name == "technician_performance":
                self._export_performance_to_csv(writer, data)
            elif report_name == "ticket_volume":
                self._export_volume_to_csv(writer, data)
            elif report_name == "category_breakdown":
                self._export_category_to_csv(writer, data)
            else:
                self._export_generic_to_csv(writer, data)

        return filepath

    # ============================================================================
    # PDF HEADER/FOOTER
    # ============================================================================

    def _create_header(self, canvas, doc, report_title: str):
        """Add professional branded header"""
        canvas.saveState()

        # Purple header bar with gradient effect
        canvas.setFillColor(SUPPORTX_PURPLE)
        canvas.rect(0, doc.pagesize[1] - 60, doc.pagesize[0], 60, fill=1, stroke=0)

        # Darker accent line
        canvas.setFillColor(SUPPORTX_PURPLE_DARK)
        canvas.rect(0, doc.pagesize[1] - 64, doc.pagesize[0], 4, fill=1, stroke=0)

        # Logo - properly sized without compression
        logo_x = 25
        title_x = logo_x
        if os.path.exists(self.logo_path):
            try:
                # Draw logo with proper aspect ratio - larger size
                canvas.drawImage(
                    self.logo_path,
                    logo_x,
                    doc.pagesize[1] - 52,
                    width=44,
                    height=44,
                    preserveAspectRatio=True,
                    mask='auto'
                )
                title_x = logo_x + 55
            except Exception:
                pass

        # Report title only (no company name text next to logo)
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 18)
        canvas.drawString(title_x, doc.pagesize[1] - 30, report_title)

        # Subtitle with date on the right side
        canvas.setFont("Helvetica", 9)
        date_text = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        canvas.drawRightString(doc.pagesize[0] - 25, doc.pagesize[1] - 25, date_text)

        canvas.restoreState()

    def _create_footer(self, canvas, doc):
        """Add professional footer"""
        canvas.saveState()

        # Footer background
        canvas.setFillColor(SUPPORTX_PURPLE_LIGHT)
        canvas.rect(0, 0, doc.pagesize[0], 25, fill=1, stroke=0)

        # Accent line
        canvas.setFillColor(SUPPORTX_PURPLE)
        canvas.rect(0, 25, doc.pagesize[0], 2, fill=1, stroke=0)

        # Page number
        canvas.setFillColor(SUPPORTX_GRAY)
        canvas.setFont("Helvetica", 9)
        canvas.drawCentredString(doc.pagesize[0] / 2, 9, f"Page {doc.page}")

        # Confidential notice
        canvas.setFont("Helvetica-Oblique", 8)
        canvas.drawString(40, 9, "Confidential Report")

        canvas.restoreState()

    # ============================================================================
    # CHART CREATION HELPERS
    # ============================================================================

    def _create_pie_chart(self, data: List[tuple], width: int = 300, height: int = 200, title: str = "") -> Drawing:
        """Create a professional pie chart"""
        drawing = Drawing(width, height)

        pie = Pie()
        pie.x = 80
        pie.y = 30
        pie.width = 120
        pie.height = 120
        pie.data = [d[1] for d in data if d[1] > 0]
        pie.labels = [d[0] for d in data if d[1] > 0]

        # Colors
        for i in range(len(pie.data)):
            pie.slices[i].fillColor = CHART_COLORS[i % len(CHART_COLORS)]
            pie.slices[i].strokeColor = colors.white
            pie.slices[i].strokeWidth = 2

        pie.slices.strokeWidth = 1
        pie.slices.fontName = 'Helvetica'
        pie.slices.fontSize = 8

        drawing.add(pie)

        # Legend
        legend = Legend()
        legend.x = 220
        legend.y = height - 50
        legend.dx = 8
        legend.dy = 8
        legend.fontName = 'Helvetica'
        legend.fontSize = 8
        legend.boxAnchor = 'nw'
        legend.columnMaximum = 8
        legend.strokeWidth = 0.5
        legend.strokeColor = colors.HexColor('#E5E7EB')
        legend.deltax = 75
        legend.deltay = 12
        legend.autoXPadding = 5
        legend.yGap = 0
        legend.dxTextSpace = 5
        legend.alignment = 'right'
        legend.dividerLines = 1 | 2 | 4
        legend.dividerOffsY = 4.5
        legend.subCols.rpad = 30

        legend_data = [(CHART_COLORS[i % len(CHART_COLORS)], f"{data[i][0]}: {data[i][1]}") for i in range(len(data)) if data[i][1] > 0]
        legend.colorNamePairs = legend_data

        drawing.add(legend)

        # Title
        if title:
            drawing.add(String(width / 2, height - 10, title, fontSize=11, fontName='Helvetica-Bold', fillColor=SUPPORTX_PURPLE_DARK, textAnchor='middle'))

        return drawing

    def _create_bar_chart(self, labels: List[str], data_series: List[tuple], width: int = 450, height: int = 200, title: str = "") -> Drawing:
        """Create a professional bar chart with multiple series"""
        drawing = Drawing(width, height)

        bc = VerticalBarChart()
        bc.x = 60
        bc.y = 40
        bc.height = height - 80
        bc.width = width - 120
        bc.data = [series[1] for series in data_series]
        bc.categoryAxis.categoryNames = labels

        # Styling
        bc.valueAxis.valueMin = 0
        bc.valueAxis.labels.fontName = 'Helvetica'
        bc.valueAxis.labels.fontSize = 8
        bc.categoryAxis.labels.fontName = 'Helvetica'
        bc.categoryAxis.labels.fontSize = 8
        bc.categoryAxis.labels.angle = 30
        bc.categoryAxis.labels.boxAnchor = 'ne'
        bc.barWidth = 15
        bc.groupSpacing = 10

        # Colors for each series
        for i, series in enumerate(data_series):
            bc.bars[i].fillColor = CHART_COLORS[i % len(CHART_COLORS)]
            bc.bars[i].strokeColor = None

        drawing.add(bc)

        # Legend
        if len(data_series) > 1:
            legend = Legend()
            legend.x = width - 100
            legend.y = height - 30
            legend.fontName = 'Helvetica'
            legend.fontSize = 8
            legend.colorNamePairs = [(CHART_COLORS[i % len(CHART_COLORS)], series[0]) for i, series in enumerate(data_series)]
            drawing.add(legend)

        # Title
        if title:
            drawing.add(String(width / 2, height - 10, title, fontSize=11, fontName='Helvetica-Bold', fillColor=SUPPORTX_PURPLE_DARK, textAnchor='middle'))

        return drawing

    def _create_horizontal_bar_chart(self, labels: List[str], values: List[float], width: int = 450, height: int = 200, title: str = "", color=None) -> Drawing:
        """Create a horizontal bar chart for single series"""
        drawing = Drawing(width, height)

        # Create bars manually for horizontal effect
        bar_height = min(20, (height - 60) / len(labels) - 5)
        max_val = max(values) if values else 1
        bar_area_width = width - 150

        y_start = height - 50
        for i, (label, value) in enumerate(zip(labels, values)):
            y = y_start - i * (bar_height + 8)

            # Label
            drawing.add(String(5, y + bar_height / 2 - 3, label[:15], fontSize=8, fontName='Helvetica', fillColor=SUPPORTX_GRAY))

            # Bar background
            drawing.add(Rect(100, y, bar_area_width, bar_height, fillColor=colors.HexColor('#F3F4F6'), strokeColor=None))

            # Bar value
            bar_width = (value / max_val) * bar_area_width if max_val > 0 else 0
            fill_color = color or CHART_COLORS[i % len(CHART_COLORS)]
            drawing.add(Rect(100, y, bar_width, bar_height, fillColor=fill_color, strokeColor=None))

            # Value label
            drawing.add(String(105 + bar_width, y + bar_height / 2 - 3, f"{value:.1f}%" if isinstance(value, float) else str(value), fontSize=8, fontName='Helvetica-Bold', fillColor=SUPPORTX_PURPLE_DARK))

        # Title
        if title:
            drawing.add(String(width / 2, height - 10, title, fontSize=11, fontName='Helvetica-Bold', fillColor=SUPPORTX_PURPLE_DARK, textAnchor='middle'))

        return drawing

    def _create_metric_card(self, label: str, value: str, width: int = 120, height: int = 60, color=SUPPORTX_PURPLE) -> Drawing:
        """Create a metric card for KPIs"""
        drawing = Drawing(width, height)

        # Card background
        drawing.add(Rect(0, 0, width, height, fillColor=colors.white, strokeColor=colors.HexColor('#E5E7EB'), strokeWidth=1, rx=5, ry=5))

        # Accent bar
        drawing.add(Rect(0, height - 5, width, 5, fillColor=color, strokeColor=None, rx=5, ry=5))

        # Value
        drawing.add(String(width / 2, height / 2 + 5, str(value), fontSize=18, fontName='Helvetica-Bold', fillColor=SUPPORTX_PURPLE_DARK, textAnchor='middle'))

        # Label
        drawing.add(String(width / 2, 12, label, fontSize=8, fontName='Helvetica', fillColor=SUPPORTX_GRAY, textAnchor='middle'))

        return drawing

    # ============================================================================
    # SLA COMPLIANCE EXPORTS
    # ============================================================================

    def _export_sla_to_excel(self, wb, data, header_fill, header_font, title_font, subtitle_font, alt_fill, border):
        ws = wb.active
        ws.title = "SLA Summary"

        # Title section
        ws['A1'] = "SLA Compliance Report"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='left')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30

        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A2'].font = subtitle_font
        ws.merge_cells('A2:F2')

        # Summary metrics in cards style
        row = 4
        metrics = [
            ("Overall Compliance", f"{data.get('overall_compliance', 0):.1f}%"),
            ("Response Compliance", f"{data.get('response_compliance', 0):.1f}%"),
            ("Resolution Compliance", f"{data.get('resolution_compliance', 0):.1f}%"),
            ("Total Tickets", str(data.get('total_tickets', 0))),
            ("At Risk", str(data.get('at_risk_count', 0))),
            ("Breached", str(data.get('breached_count', 0))),
        ]

        for col, (metric, value) in enumerate(metrics, 1):
            cell = ws.cell(row=row, column=col, value=metric)
            cell.font = Font(bold=True, size=9, color="6B7280")
            cell.alignment = Alignment(horizontal='center')

            value_cell = ws.cell(row=row + 1, column=col, value=value)
            value_cell.font = Font(bold=True, size=14, color="6B21A8")
            value_cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Priority breakdown with chart
        if data.get('by_priority'):
            row = 8
            ws.cell(row=row, column=1, value="Compliance by Priority").font = subtitle_font
            ws.merge_cells(f'A{row}:E{row}')

            row += 1
            headers = ['Priority', 'Total', 'Met', 'Breached', 'Compliance %']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            for i, item in enumerate(data['by_priority'], start=row + 1):
                ws.cell(row=i, column=1, value=item.get('priority', '')).border = border
                ws.cell(row=i, column=2, value=item.get('total', 0)).border = border
                ws.cell(row=i, column=3, value=item.get('met', 0)).border = border
                ws.cell(row=i, column=4, value=item.get('breached', 0)).border = border
                compliance_cell = ws.cell(row=i, column=5, value=f"{item.get('compliance', 0):.1f}%")
                compliance_cell.border = border
                # Color code compliance
                if item.get('compliance', 0) >= 90:
                    compliance_cell.fill = PatternFill(start_color=EXCEL_GREEN, end_color=EXCEL_GREEN, fill_type="solid")
                    compliance_cell.font = Font(color="FFFFFF", bold=True)
                elif item.get('compliance', 0) < 80:
                    compliance_cell.fill = PatternFill(start_color=EXCEL_RED, end_color=EXCEL_RED, fill_type="solid")
                    compliance_cell.font = Font(color="FFFFFF", bold=True)

            # Add pie chart for priority breakdown
            chart_row = row + len(data['by_priority']) + 3
            chart = PieChart()
            chart.title = "SLA by Priority"
            labels = Reference(ws, min_col=1, min_row=row + 1, max_row=row + len(data['by_priority']))
            chart_data = Reference(ws, min_col=2, min_row=row, max_row=row + len(data['by_priority']))
            chart.add_data(chart_data, titles_from_data=True)
            chart.set_categories(labels)
            chart.width = 12
            chart.height = 8
            ws.add_chart(chart, f"G{row}")

    def _build_sla_pdf(self, data, title):
        styles = getSampleStyleSheet()
        story = []

        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=14, textColor=SUPPORTX_PURPLE_DARK, spaceBefore=20, spaceAfter=12, fontName='Helvetica-Bold')
        subsection_style = ParagraphStyle('SubSection', parent=styles['Heading3'], fontSize=11, textColor=SUPPORTX_GRAY, spaceBefore=15, spaceAfter=8)

        # Executive Summary with KPI Cards
        story.append(Paragraph("Executive Summary", section_style))
        story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

        # Create KPI cards row
        kpi_data = [
            [
                self._create_metric_card("Overall Compliance", f"{data.get('overall_compliance', 0):.1f}%", color=SUPPORTX_PURPLE),
                self._create_metric_card("Response SLA", f"{data.get('response_compliance', 0):.1f}%", color=SUPPORTX_BLUE),
                self._create_metric_card("Resolution SLA", f"{data.get('resolution_compliance', 0):.1f}%", color=SUPPORTX_GREEN),
                self._create_metric_card("Total Tickets", str(data.get('total_tickets', 0)), color=SUPPORTX_YELLOW),
            ]
        ]
        kpi_table = Table(kpi_data, colWidths=[130, 130, 130, 130])
        kpi_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 0.3*inch))

        # Risk Summary
        risk_data = [
            [
                self._create_metric_card("At Risk", str(data.get('at_risk_count', 0)), color=SUPPORTX_YELLOW),
                self._create_metric_card("Breached", str(data.get('breached_count', 0)), color=SUPPORTX_RED),
            ]
        ]
        risk_table = Table(risk_data, colWidths=[130, 130])
        risk_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(risk_table)
        story.append(Spacer(1, 0.4*inch))

        # Priority Breakdown with Chart
        if data.get('by_priority') and len(data['by_priority']) > 0:
            story.append(Paragraph("SLA Compliance by Priority", section_style))
            story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

            # Pie chart
            pie_data = [(item.get('priority', ''), item.get('total', 0)) for item in data['by_priority']]
            if any(d[1] > 0 for d in pie_data):
                chart = self._create_pie_chart(pie_data, width=500, height=180, title="Ticket Distribution by Priority")
                story.append(chart)
                story.append(Spacer(1, 0.2*inch))

            # Table
            priority_data = [['Priority', 'Total', 'Met', 'Breached', 'Compliance']]
            for item in data['by_priority']:
                compliance = item.get('compliance', 0)
                compliance_str = f"{compliance:.1f}%"
                priority_data.append([
                    item.get('priority', ''),
                    str(item.get('total', 0)),
                    str(item.get('met', 0)),
                    str(item.get('breached', 0)),
                    compliance_str
                ])

            table = Table(priority_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1*inch, 1.2*inch])
            table.setStyle(self._get_table_style())
            story.append(table)
            story.append(Spacer(1, 0.3*inch))

        # Category Breakdown
        if data.get('by_category') and len(data['by_category']) > 0:
            story.append(Paragraph("SLA Compliance by Category", section_style))
            story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

            # Horizontal bar chart for categories
            cat_labels = [item.get('category_name', '')[:15] for item in data['by_category'][:8]]
            cat_values = [item.get('compliance', 0) for item in data['by_category'][:8]]
            if cat_labels:
                chart = self._create_horizontal_bar_chart(cat_labels, cat_values, width=500, height=180, title="Compliance Rate by Category", color=SUPPORTX_PURPLE)
                story.append(chart)
                story.append(Spacer(1, 0.2*inch))

            # Table
            cat_data = [['Category', 'Total', 'Met', 'Breached', 'Compliance']]
            for item in data['by_category'][:10]:
                cat_data.append([
                    item.get('category_name', '')[:25],
                    str(item.get('total', 0)),
                    str(item.get('met', 0)),
                    str(item.get('breached', 0)),
                    f"{item.get('compliance', 0):.1f}%"
                ])

            table = Table(cat_data, colWidths=[2*inch, 0.9*inch, 0.9*inch, 0.9*inch, 1.1*inch])
            table.setStyle(self._get_table_style())
            story.append(table)

        # Trend Analysis
        if data.get('trend_data') and len(data['trend_data']) > 0:
            story.append(PageBreak())
            story.append(Paragraph("SLA Trend Analysis", section_style))
            story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

            # Bar chart for trends
            trend_labels = [item.get('date', '')[-5:] for item in data['trend_data'][-14:]]
            met_values = [item.get('met', 0) for item in data['trend_data'][-14:]]
            breached_values = [item.get('breached', 0) for item in data['trend_data'][-14:]]

            if trend_labels:
                chart = self._create_bar_chart(
                    trend_labels,
                    [('Met', met_values), ('Breached', breached_values)],
                    width=500, height=200, title="Daily SLA Performance"
                )
                story.append(chart)

        return story

    def _export_sla_to_csv(self, writer, data):
        writer.writerow(['SLA Compliance Report'])
        writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'])
        writer.writerow([])

        writer.writerow(['EXECUTIVE SUMMARY'])
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Overall Compliance', f"{data.get('overall_compliance', 0):.1f}%"])
        writer.writerow(['Response Compliance', f"{data.get('response_compliance', 0):.1f}%"])
        writer.writerow(['Resolution Compliance', f"{data.get('resolution_compliance', 0):.1f}%"])
        writer.writerow(['Total Tickets', data.get('total_tickets', 0)])
        writer.writerow(['At Risk Count', data.get('at_risk_count', 0)])
        writer.writerow(['Breached Count', data.get('breached_count', 0)])
        writer.writerow([])

        if data.get('by_priority'):
            writer.writerow(['PRIORITY BREAKDOWN'])
            writer.writerow(['Priority', 'Total', 'Met', 'Breached', 'Compliance %'])
            for item in data['by_priority']:
                writer.writerow([item.get('priority'), item.get('total'), item.get('met'), item.get('breached'), f"{item.get('compliance', 0):.1f}%"])
            writer.writerow([])

        if data.get('by_category'):
            writer.writerow(['CATEGORY BREAKDOWN'])
            writer.writerow(['Category', 'Total', 'Met', 'Breached', 'Compliance %'])
            for item in data['by_category']:
                writer.writerow([item.get('category_name'), item.get('total'), item.get('met'), item.get('breached'), f"{item.get('compliance', 0):.1f}%"])

    # ============================================================================
    # TICKET AGING EXPORTS
    # ============================================================================

    def _export_aging_to_excel(self, wb, data, header_fill, header_font, title_font, subtitle_font, alt_fill, border):
        ws = wb.active
        ws.title = "Aging Summary"

        ws['A1'] = "Ticket Aging Report"
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30

        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A2'].font = subtitle_font
        ws.merge_cells('A2:F2')

        # Total open tickets highlight
        ws['A4'] = "Total Open Tickets"
        ws['A4'].font = Font(bold=True, size=10, color="6B7280")
        ws['B4'] = data.get('total_open_tickets', 0)
        ws['B4'].font = Font(bold=True, size=16, color="6B21A8")

        # Aging buckets
        row = 6
        ws.cell(row=row, column=1, value="Aging Distribution").font = subtitle_font

        row += 1
        headers = ['Age Bucket', 'Ticket Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        summary = data.get('summary', {})
        total = sum(summary.values()) if summary else 1
        for i, (bucket, count) in enumerate(summary.items(), start=row + 1):
            ws.cell(row=i, column=1, value=bucket).border = border
            ws.cell(row=i, column=2, value=count).border = border
            pct = (count / total * 100) if total > 0 else 0
            ws.cell(row=i, column=3, value=f"{pct:.1f}%").border = border

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        # Add chart
        if summary:
            chart = BarChart()
            chart.title = "Tickets by Age"
            chart.type = "col"
            data_ref = Reference(ws, min_col=2, min_row=row, max_row=row + len(summary))
            cats = Reference(ws, min_col=1, min_row=row + 1, max_row=row + len(summary))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 12
            chart.height = 8
            ws.add_chart(chart, "E6")

        # Ticket details sheet
        tickets_ws = wb.create_sheet("Ticket Details")
        headers = ['Ticket #', 'Title', 'Priority', 'Status', 'Assignee', 'Age (Days)', 'Age (Hours)']
        for col, header in enumerate(headers, 1):
            cell = tickets_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row_idx = 2
        buckets = data.get('buckets', {})
        for bucket_name, bucket_data in buckets.items():
            for ticket in bucket_data.get('tickets', []):
                tickets_ws.cell(row=row_idx, column=1, value=ticket.get('ticket_number', ''))
                tickets_ws.cell(row=row_idx, column=2, value=ticket.get('title', '')[:50])
                tickets_ws.cell(row=row_idx, column=3, value=ticket.get('priority', ''))
                tickets_ws.cell(row=row_idx, column=4, value=ticket.get('status', ''))
                tickets_ws.cell(row=row_idx, column=5, value=ticket.get('assignee_name', ''))
                tickets_ws.cell(row=row_idx, column=6, value=ticket.get('age_days', 0))
                tickets_ws.cell(row=row_idx, column=7, value=ticket.get('age_hours', 0))
                row_idx += 1

    def _build_aging_pdf(self, data, title):
        styles = getSampleStyleSheet()
        story = []

        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=14, textColor=SUPPORTX_PURPLE_DARK, spaceBefore=20, spaceAfter=12, fontName='Helvetica-Bold')

        # Summary Header
        story.append(Paragraph("Aging Overview", section_style))
        story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

        # KPI Card for total
        kpi_data = [[self._create_metric_card("Total Open Tickets", str(data.get('total_open_tickets', 0)), width=150, height=70, color=SUPPORTX_PURPLE)]]
        kpi_table = Table(kpi_data)
        story.append(kpi_table)
        story.append(Spacer(1, 0.3*inch))

        # Aging buckets chart
        summary = data.get('summary', {})
        if summary:
            # Pie chart
            pie_data = [(bucket, count) for bucket, count in summary.items()]
            chart = self._create_pie_chart(pie_data, width=500, height=200, title="Ticket Age Distribution")
            story.append(chart)
            story.append(Spacer(1, 0.3*inch))

            # Table
            story.append(Paragraph("Aging Breakdown", section_style))
            story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

            total = sum(summary.values()) if summary else 1
            bucket_data = [['Age Bucket', 'Count', 'Percentage']]
            for bucket, count in summary.items():
                pct = (count / total * 100) if total > 0 else 0
                bucket_data.append([bucket, str(count), f"{pct:.1f}%"])

            table = Table(bucket_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
            table.setStyle(self._get_table_style())
            story.append(table)
            story.append(Spacer(1, 0.3*inch))

        # Sample tickets
        story.append(Paragraph("Critical Aging Tickets (Sample)", section_style))
        story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

        ticket_data = [['Ticket #', 'Priority', 'Status', 'Assignee', 'Age']]
        count = 0
        buckets = data.get('buckets', {})
        # Prioritize older tickets
        for bucket_name in ['30+ days', '14-30 days', '7-14 days', '3-7 days', '1-3 days', '0-24 hours']:
            if bucket_name in buckets:
                for ticket in buckets[bucket_name].get('tickets', [])[:5]:
                    if count >= 15:
                        break
                    ticket_data.append([
                        ticket.get('ticket_number', ''),
                        ticket.get('priority', ''),
                        ticket.get('status', ''),
                        ticket.get('assignee_name', '')[:15],
                        f"{ticket.get('age_days', 0)}d {ticket.get('age_hours', 0) % 24}h"
                    ])
                    count += 1

        if len(ticket_data) > 1:
            table = Table(ticket_data, colWidths=[1.2*inch, 1*inch, 1.2*inch, 1.5*inch, 1*inch])
            table.setStyle(self._get_table_style())
            story.append(table)

        return story

    def _export_aging_to_csv(self, writer, data):
        writer.writerow(['Ticket Aging Report'])
        writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'])
        writer.writerow([])

        writer.writerow(['Total Open Tickets', data.get('total_open_tickets', 0)])
        writer.writerow([])

        writer.writerow(['AGING SUMMARY'])
        writer.writerow(['Age Bucket', 'Count'])
        summary = data.get('summary', {})
        for bucket, count in summary.items():
            writer.writerow([bucket, count])
        writer.writerow([])

        writer.writerow(['TICKET DETAILS'])
        writer.writerow(['Ticket #', 'Title', 'Priority', 'Status', 'Assignee', 'Age (Days)', 'Age (Hours)'])
        buckets = data.get('buckets', {})
        for bucket_name, bucket_data in buckets.items():
            for ticket in bucket_data.get('tickets', []):
                writer.writerow([
                    ticket.get('ticket_number', ''),
                    ticket.get('title', ''),
                    ticket.get('priority', ''),
                    ticket.get('status', ''),
                    ticket.get('assignee_name', ''),
                    ticket.get('age_days', 0),
                    ticket.get('age_hours', 0)
                ])

    # ============================================================================
    # TECHNICIAN PERFORMANCE EXPORTS
    # ============================================================================

    def _export_performance_to_excel(self, wb, data, header_fill, header_font, title_font, subtitle_font, alt_fill, border):
        ws = wb.active
        ws.title = "Team Performance"

        ws['A1'] = "Team Performance Report"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 30

        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A2'].font = subtitle_font
        ws.merge_cells('A2:H2')

        # Team summary
        if isinstance(data, list) and len(data) > 0:
            total_tickets = sum(t.get('total_tickets', 0) for t in data)
            total_resolved = sum(t.get('resolved_tickets', 0) for t in data)
            avg_sla = sum(t.get('sla_compliance', 0) for t in data) / len(data)

            ws['A4'] = "Team Members"
            ws['A4'].font = Font(bold=True, size=9, color="6B7280")
            ws['B4'] = len(data)
            ws['B4'].font = Font(bold=True, size=14, color="6B21A8")

            ws['C4'] = "Total Tickets"
            ws['C4'].font = Font(bold=True, size=9, color="6B7280")
            ws['D4'] = total_tickets
            ws['D4'].font = Font(bold=True, size=14, color="6B21A8")

            ws['E4'] = "Avg SLA"
            ws['E4'].font = Font(bold=True, size=9, color="6B7280")
            ws['F4'] = f"{avg_sla:.1f}%"
            ws['F4'].font = Font(bold=True, size=14, color="6B21A8")

        # Performance table
        row = 7
        headers = ['Technician', 'Email', 'Total', 'Resolved', 'Open', 'Resolution %', 'Avg Time (h)', 'SLA %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 15

        if isinstance(data, list):
            for i, tech in enumerate(data, start=row + 1):
                ws.cell(row=i, column=1, value=tech.get('user_name', '')).border = border
                ws.cell(row=i, column=2, value=tech.get('email', '')).border = border
                ws.cell(row=i, column=3, value=tech.get('total_tickets', 0)).border = border
                ws.cell(row=i, column=4, value=tech.get('resolved_tickets', 0)).border = border
                ws.cell(row=i, column=5, value=tech.get('open_tickets', 0)).border = border
                ws.cell(row=i, column=6, value=f"{tech.get('resolution_rate', 0):.1f}%").border = border
                ws.cell(row=i, column=7, value=f"{tech.get('avg_resolution_time_hours', 0):.1f}").border = border

                sla_cell = ws.cell(row=i, column=8, value=f"{tech.get('sla_compliance', 0):.1f}%")
                sla_cell.border = border
                if tech.get('sla_compliance', 0) >= 90:
                    sla_cell.fill = PatternFill(start_color=EXCEL_GREEN, end_color=EXCEL_GREEN, fill_type="solid")
                    sla_cell.font = Font(color="FFFFFF", bold=True)
                elif tech.get('sla_compliance', 0) < 80:
                    sla_cell.fill = PatternFill(start_color=EXCEL_RED, end_color=EXCEL_RED, fill_type="solid")
                    sla_cell.font = Font(color="FFFFFF", bold=True)

            # Add bar chart for tickets
            if len(data) > 0:
                chart = BarChart()
                chart.title = "Tickets by Team Member"
                chart.type = "col"
                data_ref = Reference(ws, min_col=3, min_row=row, max_row=row + len(data), max_col=4)
                cats = Reference(ws, min_col=1, min_row=row + 1, max_row=row + len(data))
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)
                chart.width = 14
                chart.height = 8
                ws.add_chart(chart, "J7")

    def _build_performance_pdf(self, data, title):
        styles = getSampleStyleSheet()
        story = []

        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=14, textColor=SUPPORTX_PURPLE_DARK, spaceBefore=20, spaceAfter=12, fontName='Helvetica-Bold')

        story.append(Paragraph("Team Performance Overview", section_style))
        story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

        if isinstance(data, list) and len(data) > 0:
            # Team KPIs
            total_tickets = sum(t.get('total_tickets', 0) for t in data)
            total_resolved = sum(t.get('resolved_tickets', 0) for t in data)
            avg_sla = sum(t.get('sla_compliance', 0) for t in data) / len(data)
            avg_resolution = sum(t.get('avg_resolution_time_hours', 0) for t in data) / len(data)

            kpi_data = [
                [
                    self._create_metric_card("Team Size", str(len(data)), color=SUPPORTX_PURPLE),
                    self._create_metric_card("Total Tickets", str(total_tickets), color=SUPPORTX_BLUE),
                    self._create_metric_card("Resolved", str(total_resolved), color=SUPPORTX_GREEN),
                    self._create_metric_card("Avg SLA", f"{avg_sla:.1f}%", color=SUPPORTX_YELLOW),
                ]
            ]
            kpi_table = Table(kpi_data, colWidths=[130, 130, 130, 130])
            story.append(kpi_table)
            story.append(Spacer(1, 0.4*inch))

            # Performance bar chart
            tech_names = [t.get('user_name', '')[:12] for t in data[:10]]
            tickets_resolved = [t.get('resolved_tickets', 0) for t in data[:10]]
            tickets_open = [t.get('open_tickets', 0) for t in data[:10]]

            if tech_names:
                chart = self._create_bar_chart(
                    tech_names,
                    [('Resolved', tickets_resolved), ('Open', tickets_open)],
                    width=500, height=200, title="Tickets by Team Member"
                )
                story.append(chart)
                story.append(Spacer(1, 0.3*inch))

            # SLA Compliance chart
            sla_values = [t.get('sla_compliance', 0) for t in data[:10]]
            chart = self._create_horizontal_bar_chart(tech_names, sla_values, width=500, height=200, title="SLA Compliance by Team Member")
            story.append(chart)
            story.append(Spacer(1, 0.3*inch))

            # Detailed table
            story.append(Paragraph("Individual Performance Details", section_style))
            story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

            perf_data = [['Name', 'Tickets', 'Resolved', 'Resolution %', 'Avg Time', 'SLA']]
            for tech in data:
                perf_data.append([
                    tech.get('user_name', '')[:18],
                    str(tech.get('total_tickets', 0)),
                    str(tech.get('resolved_tickets', 0)),
                    f"{tech.get('resolution_rate', 0):.1f}%",
                    f"{tech.get('avg_resolution_time_hours', 0):.1f}h",
                    f"{tech.get('sla_compliance', 0):.1f}%"
                ])

            table = Table(perf_data, colWidths=[1.6*inch, 0.8*inch, 0.9*inch, 1*inch, 0.9*inch, 0.8*inch])
            table.setStyle(self._get_table_style())
            story.append(table)
        else:
            story.append(Paragraph("No performance data available for this period.", styles['Normal']))

        return story

    def _export_performance_to_csv(self, writer, data):
        writer.writerow(['Team Performance Report'])
        writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'])
        writer.writerow([])

        if isinstance(data, list) and len(data) > 0:
            total_tickets = sum(t.get('total_tickets', 0) for t in data)
            total_resolved = sum(t.get('resolved_tickets', 0) for t in data)
            avg_sla = sum(t.get('sla_compliance', 0) for t in data) / len(data)

            writer.writerow(['TEAM SUMMARY'])
            writer.writerow(['Team Size', len(data)])
            writer.writerow(['Total Tickets', total_tickets])
            writer.writerow(['Total Resolved', total_resolved])
            writer.writerow(['Average SLA Compliance', f"{avg_sla:.1f}%"])
            writer.writerow([])

        writer.writerow(['INDIVIDUAL PERFORMANCE'])
        writer.writerow(['Technician', 'Email', 'Total Tickets', 'Resolved', 'Open', 'Resolution Rate', 'Avg Resolution (hrs)', 'SLA Compliance'])

        if isinstance(data, list):
            for tech in data:
                writer.writerow([
                    tech.get('user_name', ''),
                    tech.get('email', ''),
                    tech.get('total_tickets', 0),
                    tech.get('resolved_tickets', 0),
                    tech.get('open_tickets', 0),
                    f"{tech.get('resolution_rate', 0):.1f}%",
                    f"{tech.get('avg_resolution_time_hours', 0):.1f}",
                    f"{tech.get('sla_compliance', 0):.1f}%"
                ])

    # ============================================================================
    # TICKET VOLUME EXPORTS
    # ============================================================================

    def _export_volume_to_excel(self, wb, data, header_fill, header_font, title_font, subtitle_font, alt_fill, border):
        ws = wb.active
        ws.title = "Volume Trends"

        ws['A1'] = "Ticket Volume Report"
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30

        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A2'].font = subtitle_font
        ws.merge_cells('A2:F2')

        # Summary KPIs
        ws['A4'] = "Total Created"
        ws['A4'].font = Font(bold=True, size=9, color="6B7280")
        ws['B4'] = data.get('total_created', 0)
        ws['B4'].font = Font(bold=True, size=14, color="6B21A8")

        ws['C4'] = "Total Resolved"
        ws['C4'].font = Font(bold=True, size=9, color="6B7280")
        ws['D4'] = data.get('total_resolved', 0)
        ws['D4'].font = Font(bold=True, size=14, color="10B981")

        # Trend data
        row = 7
        headers = ['Period', 'Created', 'Resolved']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        trend_data = data.get('trend_data', [])
        for i, item in enumerate(trend_data, start=row + 1):
            ws.cell(row=i, column=1, value=item.get('period', '')).border = border
            ws.cell(row=i, column=2, value=item.get('created', 0)).border = border
            ws.cell(row=i, column=3, value=item.get('resolved', 0)).border = border

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12

        # Add line chart
        if trend_data:
            chart = LineChart()
            chart.title = "Ticket Volume Trend"
            chart.style = 10
            chart.y_axis.title = "Tickets"
            chart.x_axis.title = "Period"

            data_ref = Reference(ws, min_col=2, min_row=row, max_row=row + len(trend_data), max_col=3)
            cats = Reference(ws, min_col=1, min_row=row + 1, max_row=row + len(trend_data))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 14
            chart.height = 8
            ws.add_chart(chart, "E7")

    def _build_volume_pdf(self, data, title):
        styles = getSampleStyleSheet()
        story = []

        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=14, textColor=SUPPORTX_PURPLE_DARK, spaceBefore=20, spaceAfter=12, fontName='Helvetica-Bold')

        story.append(Paragraph("Volume Overview", section_style))
        story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

        # KPI Cards
        kpi_data = [
            [
                self._create_metric_card("Total Created", str(data.get('total_created', 0)), color=SUPPORTX_BLUE),
                self._create_metric_card("Total Resolved", str(data.get('total_resolved', 0)), color=SUPPORTX_GREEN),
                self._create_metric_card("Net Change", str(data.get('total_created', 0) - data.get('total_resolved', 0)), color=SUPPORTX_YELLOW),
            ]
        ]
        kpi_table = Table(kpi_data, colWidths=[150, 150, 150])
        story.append(kpi_table)
        story.append(Spacer(1, 0.4*inch))

        # Trend chart
        trend_data = data.get('trend_data', [])
        if trend_data:
            story.append(Paragraph("Volume Trend Analysis", section_style))
            story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

            labels = [item.get('period', '')[-5:] for item in trend_data[-21:]]
            created = [item.get('created', 0) for item in trend_data[-21:]]
            resolved = [item.get('resolved', 0) for item in trend_data[-21:]]

            chart = self._create_bar_chart(
                labels,
                [('Created', created), ('Resolved', resolved)],
                width=500, height=220, title="Daily Ticket Volume"
            )
            story.append(chart)
            story.append(Spacer(1, 0.3*inch))

            # Summary table
            story.append(Paragraph("Daily Breakdown", section_style))
            story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

            vol_data = [['Period', 'Created', 'Resolved', 'Net']]
            for item in trend_data[-14:]:
                net = item.get('created', 0) - item.get('resolved', 0)
                vol_data.append([
                    item.get('period', ''),
                    str(item.get('created', 0)),
                    str(item.get('resolved', 0)),
                    f"+{net}" if net > 0 else str(net)
                ])

            table = Table(vol_data, colWidths=[2*inch, 1.2*inch, 1.2*inch, 1*inch])
            table.setStyle(self._get_table_style())
            story.append(table)

        return story

    def _export_volume_to_csv(self, writer, data):
        writer.writerow(['Ticket Volume Report'])
        writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'])
        writer.writerow([])

        writer.writerow(['SUMMARY'])
        writer.writerow(['Total Created', data.get('total_created', 0)])
        writer.writerow(['Total Resolved', data.get('total_resolved', 0)])
        writer.writerow([])

        writer.writerow(['VOLUME TREND'])
        writer.writerow(['Period', 'Created', 'Resolved'])
        trend_data = data.get('trend_data', [])
        for item in trend_data:
            writer.writerow([item.get('period', ''), item.get('created', 0), item.get('resolved', 0)])

    # ============================================================================
    # CATEGORY BREAKDOWN EXPORTS
    # ============================================================================

    def _export_category_to_excel(self, wb, data, header_fill, header_font, title_font, subtitle_font, alt_fill, border):
        ws = wb.active
        ws.title = "Categories"

        ws['A1'] = "Category Breakdown Report"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')
        ws.row_dimensions[1].height = 30

        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A2'].font = subtitle_font
        ws.merge_cells('A2:G2')

        row = 4
        headers = ['Category', 'Total', 'Resolved', 'Open', 'Resolution %', 'Avg Time (h)', 'SLA %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 15

        if isinstance(data, list):
            for i, cat in enumerate(data, start=row + 1):
                ws.cell(row=i, column=1, value=cat.get('category_name', '')).border = border
                ws.cell(row=i, column=2, value=cat.get('total_tickets', 0)).border = border
                ws.cell(row=i, column=3, value=cat.get('resolved_tickets', 0)).border = border
                ws.cell(row=i, column=4, value=cat.get('open_tickets', 0)).border = border
                ws.cell(row=i, column=5, value=f"{cat.get('resolution_rate', 0):.1f}%").border = border
                ws.cell(row=i, column=6, value=f"{cat.get('avg_resolution_time_hours', 0):.1f}").border = border
                ws.cell(row=i, column=7, value=f"{cat.get('sla_compliance', 0):.1f}%").border = border

            # Add pie chart
            if len(data) > 0:
                chart = PieChart()
                chart.title = "Tickets by Category"
                labels = Reference(ws, min_col=1, min_row=row + 1, max_row=row + len(data))
                chart_data = Reference(ws, min_col=2, min_row=row, max_row=row + len(data))
                chart.add_data(chart_data, titles_from_data=True)
                chart.set_categories(labels)
                chart.width = 12
                chart.height = 8
                ws.add_chart(chart, "I4")

    def _build_category_pdf(self, data, title):
        styles = getSampleStyleSheet()
        story = []

        section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=14, textColor=SUPPORTX_PURPLE_DARK, spaceBefore=20, spaceAfter=12, fontName='Helvetica-Bold')

        story.append(Paragraph("Category Distribution", section_style))
        story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

        if isinstance(data, list) and len(data) > 0:
            # Pie chart
            pie_data = [(cat.get('category_name', '')[:15], cat.get('total_tickets', 0)) for cat in data[:8]]
            if any(d[1] > 0 for d in pie_data):
                chart = self._create_pie_chart(pie_data, width=500, height=200, title="Ticket Distribution by Category")
                story.append(chart)
                story.append(Spacer(1, 0.3*inch))

            # SLA by category bar chart
            cat_names = [cat.get('category_name', '')[:12] for cat in data[:10]]
            sla_values = [cat.get('sla_compliance', 0) for cat in data[:10]]
            chart = self._create_horizontal_bar_chart(cat_names, sla_values, width=500, height=200, title="SLA Compliance by Category")
            story.append(chart)
            story.append(Spacer(1, 0.3*inch))

            # Detailed table
            story.append(Paragraph("Category Performance Details", section_style))
            story.append(HRFlowable(width="100%", thickness=2, color=SUPPORTX_PURPLE, spaceAfter=15))

            cat_data = [['Category', 'Total', 'Resolved', 'Resolution %', 'SLA']]
            for cat in data:
                cat_data.append([
                    cat.get('category_name', '')[:22],
                    str(cat.get('total_tickets', 0)),
                    str(cat.get('resolved_tickets', 0)),
                    f"{cat.get('resolution_rate', 0):.1f}%",
                    f"{cat.get('sla_compliance', 0):.1f}%"
                ])

            table = Table(cat_data, colWidths=[2*inch, 0.9*inch, 0.9*inch, 1.1*inch, 0.9*inch])
            table.setStyle(self._get_table_style())
            story.append(table)
        else:
            story.append(Paragraph("No category data available.", styles['Normal']))

        return story

    def _export_category_to_csv(self, writer, data):
        writer.writerow(['Category Breakdown Report'])
        writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'])
        writer.writerow([])

        writer.writerow(['CATEGORY PERFORMANCE'])
        writer.writerow(['Category', 'Total Tickets', 'Resolved', 'Open', 'Resolution Rate', 'Avg Resolution (hrs)', 'SLA Compliance'])

        if isinstance(data, list):
            for cat in data:
                writer.writerow([
                    cat.get('category_name', ''),
                    cat.get('total_tickets', 0),
                    cat.get('resolved_tickets', 0),
                    cat.get('open_tickets', 0),
                    f"{cat.get('resolution_rate', 0):.1f}%",
                    f"{cat.get('avg_resolution_time_hours', 0):.1f}",
                    f"{cat.get('sla_compliance', 0):.1f}%"
                ])

    # ============================================================================
    # GENERIC EXPORTS
    # ============================================================================

    def _export_generic_to_excel(self, ws, data, header_fill, header_font, alt_fill, border):
        if isinstance(data, dict):
            row = 1
            for key, value in data.items():
                ws.cell(row=row, column=1, value=str(key))
                ws.cell(row=row, column=2, value=str(value))
                row += 1
        elif isinstance(data, list) and len(data) > 0:
            if isinstance(data[0], dict):
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill

                for row_idx, item in enumerate(data, 2):
                    for col, key in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col, value=str(item.get(key, '')))

    def _build_generic_pdf(self, data, title):
        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph("Report exported successfully.", styles['Normal']))
        return story

    def _export_generic_to_csv(self, writer, data):
        if isinstance(data, dict):
            for key, value in data.items():
                writer.writerow([str(key), str(value)])
        elif isinstance(data, list) and len(data) > 0:
            if isinstance(data[0], dict):
                headers = list(data[0].keys())
                writer.writerow(headers)
                for item in data:
                    writer.writerow([str(item.get(k, '')) for k in headers])

    # ============================================================================
    # HELPER METHODS
    # ============================================================================

    def _get_table_style(self):
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), SUPPORTX_PURPLE),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, SUPPORTX_PURPLE_LIGHT]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ])

    def cleanup_old_exports(self, days: int = 7):
        """Remove export files older than specified days"""
        import time
        current_time = time.time()

        for filename in os.listdir(self.export_dir):
            filepath = os.path.join(self.export_dir, filename)
            if os.path.isfile(filepath):
                file_age = current_time - os.path.getmtime(filepath)
                if file_age > days * 86400:
                    os.remove(filepath)
